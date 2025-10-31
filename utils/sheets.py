"""
Модуль для работы с Excel файлами для сохранения вопросов и отзывов
"""
import os
import asyncio
import fcntl  # для блокировок файлов на Linux/Unix
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from config import EXCEL_FILE_PATH


class ExcelManager:
    def __init__(self):
        self.file_path = EXCEL_FILE_PATH
        self._excel_lock = asyncio.Lock()  # Блокировка для синхронизации доступа к Excel
        self._init_file()
    
    def _init_file(self):
        """Инициализация Excel файла с листами"""
        if not os.path.exists(self.file_path):
            # Создаем новый файл
            wb = Workbook()
            wb.remove(wb.active)  # Удаляем дефолтный лист
            
            # Создаем лист "Вопросы"
            ws_questions = wb.create_sheet("Вопросы")
            ws_questions.append(["ID пользователя", "Имя", "Вопрос", "Дата"])
            self._format_header(ws_questions)
            
            # Создаем лист "Отзывы"
            ws_feedback = wb.create_sheet("Отзывы")
            ws_feedback.append([
                "ID пользователя", "Имя", "Польза форума", "Интересные направления", 
                "Предложения по улучшению", "Дата"
            ])
            self._format_header(ws_feedback)
            
            wb.save(self.file_path)
            print(f"Создан новый Excel файл: {self.file_path}")
        else:
            # Проверяем, что листы существуют
            wb = load_workbook(self.file_path)
            
            if "Вопросы" not in wb.sheetnames:
                ws_questions = wb.create_sheet("Вопросы")
                ws_questions.append(["ID пользователя", "Имя", "Вопрос", "Дата"])
                self._format_header(ws_questions)
            
            if "Отзывы" not in wb.sheetnames:
                ws_feedback = wb.create_sheet("Отзывы")
                ws_feedback.append([
                    "ID пользователя", "Имя", "Польза форума", "Интересные направления", 
                    "Предложения по улучшению", "Дата"
                ])
                self._format_header(ws_feedback)
            
            wb.save(self.file_path)
    
    def _format_header(self, worksheet):
        """Форматирование заголовка листа"""
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Автоматическая ширина столбцов
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    def save_question(self, user_id: str, user_name: str, question_text: str):
        """Сохранение вопроса в Excel"""
        try:
            wb = load_workbook(self.file_path)
            
            if "Вопросы" not in wb.sheetnames:
                ws = wb.create_sheet("Вопросы")
                ws.append(["ID пользователя", "Имя", "Вопрос", "Дата"])
                self._format_header(ws)
            else:
                ws = wb["Вопросы"]
            
            ws.append([
                str(user_id),
                user_name,
                question_text,
                datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ])
            
            wb.save(self.file_path)
            print(f"Вопрос сохранен в Excel: {self.file_path}")
            return True
        except Exception as e:
            print(f"Ошибка сохранения вопроса в Excel: {e}")
            return False
    
    async def save_feedback(self, user_id: str, user_name: str, feedback_data: dict):
        """Сохранение отзыва в Excel (асинхронная версия с блокировкой)"""
        async with self._excel_lock:
            try:
                print(f"[DEBUG] save_feedback: Начало сохранения отзыва для user_id={user_id}")
                
                # Используем временный файл для атомарной записи
                temp_file = self.file_path + '.tmp'
                
                # Загружаем существующий файл или создаем новый
                if os.path.exists(self.file_path):
                    wb = load_workbook(self.file_path)
                else:
                    wb = Workbook()
                    wb.remove(wb.active)  # Удаляем дефолтный лист

                if "Отзывы" not in wb.sheetnames:
                    ws = wb.create_sheet("Отзывы")
                    ws.append([
                        "ID пользователя", "Имя", "Польза форума", "Интересные направления", 
                        "Предложения по улучшению", "Дата"
                    ])
                    self._format_header(ws)
                    print(f"[DEBUG] Создан лист 'Отзывы' в Excel")
                else:
                    ws = wb["Отзывы"]
                    # Проверяем заголовки - если старые, обновляем их
                    header_row = ws[1]
                    if len(header_row) < 6 or header_row[2].value != "Польза форума":
                        # Старые заголовки - заменяем первую строку
                        print(f"[DEBUG] Обновление заголовков в листе 'Отзывы'")
                        # Сохраняем данные (кроме заголовка)
                        all_data = list(ws.iter_rows(values_only=True))
                        ws.delete_rows(1, ws.max_row)
                        ws.append([
                            "ID пользователя", "Имя", "Польза форума", "Интересные направления", 
                            "Предложения по улучшению", "Дата"
                        ])
                        self._format_header(ws)
                        # Восстанавливаем данные со старой структурой (если есть)
                        if len(all_data) > 1:
                            for row in all_data[1:]:
                                if row and len(row) >= 3:
                                    # Старая структура: ID, Имя, Полный отзыв, Дата
                                    # Новая структура: ID, Имя, Польза, Направления, Предложения, Дата
                                    if len(row) == 4:
                                        ws.append([
                                            row[0], row[1], "", "", row[2] if len(row) > 2 else "", row[3] if len(row) > 3 else ""
                                        ])

                # Сохраняем ответы в отдельные столбцы
                q1_benefit = feedback_data.get("q1_benefit", "")
                q2_directions = feedback_data.get("q2_directions", "")
                q3_suggestions = feedback_data.get("q3_suggestions", "")
                
                # Если есть полный отзыв (старая структура), пытаемся извлечь из него
                if not q1_benefit and not q2_directions and not q3_suggestions:
                    full_feedback = feedback_data.get("full_feedback", "")
                    if full_feedback:
                        # Старая структура - оставляем как есть, но разделяем если возможно
                        q1_benefit = full_feedback
                        q2_directions = ""
                        q3_suggestions = ""

                print(f"[DEBUG] Добавляю строку в Excel:")
                print(f"  user_id={user_id}")
                print(f"  user_name={user_name}")
                print(f"  q1_benefit={q1_benefit[:50]}...")
                print(f"  q2_directions={q2_directions[:50]}...")
                print(f"  q3_suggestions={q3_suggestions[:50]}...")
                
                ws.append([
                    str(user_id),
                    user_name,
                    q1_benefit,
                    q2_directions,
                    q3_suggestions,
                    datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                ])

                # Сохраняем во временный файл
                wb.save(temp_file)
                
                # Блокировка файла для атомарной записи (Linux/Unix)
                try:
                    with open(temp_file, 'rb+') as f:
                        fcntl.flock(f.fileno(), fcntl.LOCK_EX)
                        try:
                            # Файл уже сохранен, просто держим блокировку
                            pass
                        finally:
                            fcntl.flock(f.fileno(), fcntl.LOCK_UN)
                except (OSError, AttributeError):
                    # Если fcntl не работает (Windows), просто продолжаем
                    pass
                
                # Атомарное переименование
                os.replace(temp_file, self.file_path)
                
                print(f"[DEBUG] ✅ Отзыв успешно сохранен в Excel: {self.file_path}")
                print(f"[DEBUG] Строка добавлена в лист 'Отзывы'")
                return True
            except Exception as e:
                print(f"[DEBUG] ❌ Ошибка сохранения отзыва в Excel: {e}")
                import traceback
                traceback.print_exc()
                # Удаляем временный файл при ошибке
                temp_file = self.file_path + '.tmp'
                if os.path.exists(temp_file):
                    try:
                        os.remove(temp_file)
                    except:
                        pass
                return False


excel_manager = ExcelManager()
